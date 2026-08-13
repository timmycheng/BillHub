#!/usr/bin/env python3
"""BillHub - 合同清单批量导入导出（.xlsx）"""
import os
import re
from datetime import datetime, date

import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

import db

# 导出表头（导入时按此表头名匹配列）
EXPORT_HEADERS = ['合同编号', '合同名称', '签订单位', '经办人', '分类', '收款单位', '开户银行',
                  '银行账号', '合同总金额', '签订日期', '生效时间', '结束时间', '备注',
                  '已付金额', '剩余金额']

# 导入表头 -> db 字段（含别名容错；已付/剩余为系统计算值，导入时忽略）
HEADER_MAP = {
    '合同编号': 'contract_no',
    '合同名称': 'contract_name',
    '签订单位': 'customer_name', '客户/甲方': 'customer_name', '客户': 'customer_name', '甲方': 'customer_name',
    '经办人': 'contract_manager', '负责人': 'contract_manager', '联系人': 'contract_manager',
    '分类': 'category', '类型': 'category',
    '收款单位': 'payee', '收款方': 'payee',
    '开户银行': 'bank_name', '银行': 'bank_name',
    '银行账号': 'bank_account', '银行帐号': 'bank_account', '账号': 'bank_account',
    '合同总金额': 'total_amount', '总金额': 'total_amount', '金额': 'total_amount',
    '签订日期': 'sign_date', '签约日期': 'sign_date', '日期': 'sign_date',
    '生效时间': 'start_date', '生效日期': 'start_date', '开始时间': 'start_date', '起始日期': 'start_date',
    '结束时间': 'end_date', '结束日期': 'end_date', '到期时间': 'end_date', '到期日期': 'end_date',
    '备注': 'remark',
}

# 全角 -> 半角（数字、逗号、点、全角￥），str.strip() 不处理全角空格需单独剔除
_FULLWIDTH_MAP = str.maketrans('０１２３４５６７８９．，￥', '0123456789.,¥')


def _cell_text(value):
    """任意单元格值 -> 去首尾空格字符串；None -> ''"""
    if value is None:
        return ''
    return str(value).strip()


def _half_width(s):
    """全角字符转半角，并去除全角空格/不间断空格"""
    return s.translate(_FULLWIDTH_MAP).replace('　', '').replace(' ', '')


def parse_amount(value):
    """金额解析：数字直接返回；字符串兼容逗号/¥/￥/全角字符/空格；非法返回 None"""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return round(float(value), 2)
    s = _half_width(str(value).strip())
    if not s:
        return None
    s = s.replace('，', ',').replace('¥', '').replace('￥', '').replace(',', '')
    s = s.replace(' ', '')
    try:
        return round(float(s), 2)
    except ValueError:
        return None


def parse_date(value):
    """日期解析：datetime/date 对象 -> 'YYYY-MM-DD'；字符串兼容 2026-8-7 / 2026/08/07 /
    2026.8.7 / 2026年8月7日；空值返回 ''；非法返回 None"""
    if value is None:
        return ''
    if isinstance(value, (datetime, date)):  # openpyxl 读日期单元格返回 datetime
        return value.strftime('%Y-%m-%d')
    s = _half_width(str(value).strip())
    if not s:
        return ''
    s = s.replace('年', '-').replace('月', '-').replace('日', '')
    s = s.replace('/', '-').replace('.', '-').replace(' ', '')
    m = re.match(r'^(\d{4})-(\d{1,2})-(\d{1,2})$', s)
    if m:
        y, mo, d = (int(g) for g in m.groups())
        if 1 <= mo <= 12 and 1 <= d <= 31:
            return f'{y:04d}-{mo:02d}-{d:02d}'
    return None


def export_contracts(filepath):
    """把 contracts 全部导出为 .xlsx（表头加粗、列宽、冻结首行），返回导出条数"""
    contracts = db.list_contracts()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '合同清单'

    # 表头：加粗 + 居中
    for col, h in enumerate(EXPORT_HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    # 数据行：金额写数值（可继续编辑），日期写字符串保证往返无损
    for i, c in enumerate(contracts, start=2):
        stats = db.get_contract_stats(c['id'])
        ws.cell(row=i, column=1, value=c['contract_no'] or '')
        ws.cell(row=i, column=2, value=c['contract_name'])
        ws.cell(row=i, column=3, value=c['customer_name'] or '')
        ws.cell(row=i, column=4, value=c['contract_manager'] or '')
        ws.cell(row=i, column=5, value=c['category'] or '')
        ws.cell(row=i, column=6, value=c['payee'] or '')
        ws.cell(row=i, column=7, value=c['bank_name'] or '')
        ws.cell(row=i, column=8, value=c['bank_account'] or '')
        ws.cell(row=i, column=9, value=c['total_amount'])
        ws.cell(row=i, column=10, value=c['sign_date'] or '')
        ws.cell(row=i, column=11, value=c.get('start_date') or '')
        ws.cell(row=i, column=12, value=c.get('end_date') or '')
        ws.cell(row=i, column=13, value=c['remark'] or '')
        ws.cell(row=i, column=14, value=stats['paid'])
        ws.cell(row=i, column=15, value=stats['remaining'])
        for col in (9, 14, 15):
            ws.cell(row=i, column=col).number_format = '#,##0.00'

    # 列宽 + 冻结表头
    widths = [18, 32, 20, 10, 12, 20, 18, 20, 14, 12, 12, 12, 30, 14, 14]
    for col, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = 'A2'

    os.makedirs(os.path.dirname(filepath) or '.', exist_ok=True)
    wb.save(filepath)
    return len(contracts)


def _row_to_contract(cells, header):
    """一行单元格 -> db.add_contract 参数；返回 (参数dict, None) 或 (None, 错误原因)"""
    def get(field):
        col = header.get(field)
        return '' if col is None else cells[col].value

    # 合同编号可选（无编号合同靠隐藏 id 区分）
    name = _cell_text(get('contract_name'))
    if not name:
        return None, '合同名称不能为空'
    amount = parse_amount(get('total_amount'))
    if amount is None:
        return None, '合同总金额格式错误'
    if amount <= 0:
        return None, '合同总金额必须大于 0'
    sign_date = parse_date(get('sign_date'))
    if sign_date is None:
        return None, '签订日期格式错误'
    start_date = parse_date(get('start_date'))
    if start_date is None:
        return None, '生效时间格式错误'
    end_date = parse_date(get('end_date'))
    if end_date is None:
        return None, '结束时间格式错误'
    return {
        'contract_no': _cell_text(get('contract_no')),
        'contract_name': name,
        'customer_name': _cell_text(get('customer_name')),
        'contract_manager': _cell_text(get('contract_manager')),
        'category': _cell_text(get('category')),
        'payee': _cell_text(get('payee')),
        'bank_name': _cell_text(get('bank_name')),
        'bank_account': _cell_text(get('bank_account')),
        'total_amount': amount,
        'sign_date': sign_date,
        'start_date': start_date,
        'end_date': end_date,
        'remark': _cell_text(get('remark')),
    }, None


def import_contracts(filepath):
    """从 .xlsx 逐行导入合同。表头按列名匹配（去空格+别名容错），列序无关；
    跳过空行；编号重复的行跳过并报告。返回 {added, skipped, failed, errors}"""
    wb = openpyxl.load_workbook(filepath, data_only=True)  # 公式单元格取缓存值
    ws = wb.active

    # 表头匹配：字段名 -> 列索引(0基)
    header = {}
    for cell in ws[1]:
        name = _cell_text(cell.value)
        if name in HEADER_MAP:
            header[HEADER_MAP[name]] = cell.column - 1
    if not header:
        return {'added': 0, 'skipped': 0, 'failed': 0,
                'errors': [{'row': 1, 'reason': '第一行未找到可识别的表头'}]}

    result = {'added': 0, 'skipped': 0, 'failed': 0, 'errors': []}
    for cells in ws.iter_rows(min_row=2):
        row_no = cells[0].row  # Excel 真实行号
        if all(_cell_text(c.value) == '' for c in cells):
            continue  # 空行（含格式残留行）跳过
        data, err = _row_to_contract(cells, header)
        if err:
            result['failed'] += 1
            result['errors'].append({'row': row_no, 'reason': err})
            continue
        if db.add_contract(**data) is None:  # 编号已存在
            result['skipped'] += 1
            result['errors'].append(
                {'row': row_no, 'reason': f"合同编号「{data['contract_no']}」已存在，跳过"})
        else:
            result['added'] += 1
    return result
