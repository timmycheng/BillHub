#!/usr/bin/env python3
"""SmartBill - 模板引擎
支持 {{占位符}} 替换，格式与代码解耦。
"""
import copy
import os
import re
import openpyxl

PLACEHOLDER_RE = re.compile(r'\{\{\s*(\w+)\s*\}\}')


def render_template(template_path, output_path, context):
    """
    遍历模板 Excel 单元格，将 {{key}} 替换为 context[key]。
    保留所有格式（合并单元格、样式、边框）。
    """
    wb = openpyxl.load_workbook(template_path)
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str):
                    new_value = cell.value
                    for key, value in context.items():
                        placeholder = '{{' + key + '}}'
                        if placeholder in new_value:
                            new_value = new_value.replace(placeholder, str(value))
                    # 处理带空格的占位符 {{ key }}
                    def repl(m):
                        k = m.group(1)
                        return str(context.get(k, m.group(0)))
                    new_value = PLACEHOLDER_RE.sub(repl, new_value)
                    cell.value = new_value
    os.makedirs(os.path.dirname(output_path) or '.', exist_ok=True)
    wb.save(output_path)
    return output_path


def scan_placeholders(template_path):
    """扫描模板中使用的所有占位符（用于预览/提示）"""
    wb = openpyxl.load_workbook(template_path)
    placeholders = set()
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str):
                    for m in PLACEHOLDER_RE.finditer(cell.value):
                        placeholders.add(m.group(1))
    return sorted(placeholders)


# ============ 审批表模板渲染（付款计划行展开） ============
# 模板中付款计划数据行固定为第 10-13 行（期数 1-4），合计行 14
_PLAN_FIRST_ROW = 10
_PLAN_LAST_ROW = 13
_PLAN_SUM_ROW = 14
_PLAN_COL_KEYS = ['期数', '比例', '金额', '已支付', '本次支付', '付款依据']


def _has_plan_placeholders(ws):
    """模板是否包含付款计划占位符（防误伤旧模板）"""
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and '{{付款计划_1_' in cell.value:
                return True
    return False


def _shift_block(ws, start_row, delta):
    """把 start_row 及以下的值/样式/合并范围整体下移 delta 行。
    openpyxl 的 insert_rows 不移动合并单元格，故自写整块下移。"""
    # 1. 合并单元格重定位（只处理 start_row 及以下的区域）
    for rng in list(ws.merged_cells.ranges):
        if rng.min_row >= start_row:
            ws.unmerge_cells(str(rng))
            ws.merge_cells(start_row=rng.min_row + delta, start_column=rng.min_col,
                           end_row=rng.max_row + delta, end_column=rng.max_col)
    # 2. 值 + 样式：从下往上逐行拷贝（防覆盖）；MergedCell 只读则跳过
    max_r = ws.max_row
    for r in range(max_r, start_row - 1, -1):
        for c in range(1, ws.max_column + 1):
            src, dst = ws.cell(row=r, column=c), ws.cell(row=r + delta, column=c)
            try:
                dst.value = src.value
                dst.font = copy.copy(src.font)
                dst.border = copy.copy(src.border)
                dst.fill = copy.copy(src.fill)
                dst.alignment = copy.copy(src.alignment)
                dst.number_format = src.number_format
            except AttributeError:
                pass  # 合并区非左上角单元格，样式由左上角承载
    # 3. 新插入行样式以 start_row-1 行为源复制
    for r in range(start_row, start_row + delta):
        for c in range(1, ws.max_column + 1):
            src, dst = ws.cell(row=start_row - 1, column=c), ws.cell(row=r, column=c)
            try:
                dst.font = copy.copy(src.font)
                dst.border = copy.copy(src.border)
                dst.fill = copy.copy(src.fill)
                dst.alignment = copy.copy(src.alignment)
                dst.number_format = src.number_format
            except AttributeError:
                pass
    # 4. 行高跟随
    for r in range(start_row, max_r - delta + 1):
        ws.row_dimensions[r + delta].height = ws.row_dimensions[r].height
    # 5. 清空空出的行
    for r in range(start_row, start_row + delta):
        for c in range(1, ws.max_column + 1):
            try:
                ws.cell(row=r, column=c).value = None
            except AttributeError:
                pass
    # 4. 清空空出的行
    for r in range(start_row, start_row + delta):
        for c in range(1, ws.max_column + 1):
            try:
                ws.cell(row=r, column=c).value = None
            except AttributeError:
                pass


def _fit_plan_rows(ws, n):
    """付款计划数据行数适配：模板固定 4 行。
    n>4 → 在合计行前插入（新行样式以第 13 行为源）；n<4 → 清空多余行值（保留边框）。"""
    if not _has_plan_placeholders(ws):
        return
    if n > 4:
        _shift_block(ws, _PLAN_SUM_ROW, n - 4)
    elif n < 4:
        for r in range(_PLAN_FIRST_ROW + n, _PLAN_SUM_ROW):
            for c in range(1, ws.max_column + 1):
                ws.cell(row=r, column=c).value = None


def _replace_all(ws, context):
    """全表占位符替换。整格恰为一个数值占位符 → 直接写数值（保留数字格式）；
    否则走字符串替换（兼容混排文本）。"""
    for row in ws.iter_rows():
        for cell in row:
            v = cell.value
            if not isinstance(v, str):
                continue
            m = PLACEHOLDER_RE.fullmatch(v.strip())
            if m and m.group(1) in context and isinstance(context[m.group(1)], (int, float)):
                cell.value = context[m.group(1)]
                continue
            new_value = v
            for key, value in context.items():
                new_value = new_value.replace('{{' + key + '}}', str(value))

            def repl(m2):
                k = m2.group(1)
                return str(context.get(k, m2.group(0)))
            cell.value = PLACEHOLDER_RE.sub(repl, new_value)


def render_approval_template(template_path, output_path, context, stages):
    """审批表模板渲染：付款计划行数展开 + 占位符替换（数值保持数值型）。
    stages: db.get_plan_status 输出（含 seq/paid 等计算字段），决定行数。"""
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active
    n = len(stages)
    if _has_plan_placeholders(ws):  # 仅审批表模板执行行展开,旧模板不受影响
        _fit_plan_rows(ws, n)
        # 重建数据行占位符（序号与期数一致；n>4 时新行无占位符）
        for k in range(1, n + 1):
            row = _PLAN_FIRST_ROW + k - 1
            for col, key in enumerate(_PLAN_COL_KEYS, start=1):
                ws.cell(row=row, column=col).value = f'{{{{付款计划_{k}_{key}}}}}'
    _replace_all(ws, context)
    os.makedirs(os.path.dirname(output_path) or '.', exist_ok=True)
    wb.save(output_path)
    return output_path
