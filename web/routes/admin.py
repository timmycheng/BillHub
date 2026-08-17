"""BillHub 管理员蓝图：用户管理（增删改/重置密码/批量导入）+ OCR 规则 + 数据库备份。仅管理员可访问。"""
import io
import json
import os
import re
import sqlite3
import tempfile

import openpyxl
from flask import (Blueprint, current_app, flash, jsonify, redirect, render_template,
                   request, send_file, url_for)
from flask_login import current_user

import db
import ocr
from web.audit_log import log_action
from web.auth import admin_required, hash_password, password_error

bp = Blueprint('admin', __name__)


@bp.route('/admin/users')
@admin_required
def users():
    return render_template('admin/users.html', users=db.list_users())


@bp.route('/admin/users', methods=['POST'])
@admin_required
def users_create():
    username = request.form.get('username', '').strip()
    password = request.form.get('password', '')
    display_name = request.form.get('display_name', '').strip()
    is_admin = request.form.get('is_admin') == 'on'
    if not username:
        flash('用户名不能为空', 'danger')
        return redirect(url_for('admin.users'))
    err = password_error(password)
    if err:
        flash(err, 'danger')
        return redirect(url_for('admin.users'))
    if db.create_user(username, hash_password(password),
                      display_name=display_name, is_admin=is_admin) is None:
        flash(f'用户名「{username}」已存在', 'danger')
    else:
        flash(f'已创建用户：{username}', 'success')
        log_action('创建用户', target=username)
    return redirect(url_for('admin.users'))


@bp.route('/admin/users/<int:uid>/edit', methods=['POST'])
@admin_required
def users_edit(uid):
    user = db.get_user(uid)
    if not user:
        flash('用户不存在', 'warning')
        return redirect(url_for('admin.users'))
    display_name = request.form.get('display_name', '').strip()
    is_admin = request.form.get('is_admin') == 'on'
    password = request.form.get('password', '')
    # 防止把最后一位管理员降级成普通用户
    if user['is_admin'] and not is_admin and _admin_count() <= 1:
        flash('不能降级最后一位管理员', 'warning')
        return redirect(url_for('admin.users'))
    if password:
        err = password_error(password)
        if err:
            flash(err, 'danger')
            return redirect(url_for('admin.users'))
        db.update_user(uid, display_name=display_name, is_admin=is_admin,
                       password_hash=hash_password(password))
        flash(f'已更新用户「{user["username"]}」并重置密码', 'success')
        log_action('编辑用户', target=user['username'], detail='重置密码')
    else:
        db.update_user(uid, display_name=display_name, is_admin=is_admin)
        flash(f'已更新用户：{user["username"]}', 'success')
        log_action('编辑用户', target=user['username'])
    return redirect(url_for('admin.users'))


@bp.route('/admin/users/<int:uid>/delete', methods=['POST'])
@admin_required
def users_delete(uid):
    user = db.get_user(uid)
    if not user:
        flash('用户不存在', 'warning')
        return redirect(url_for('admin.users'))
    if uid == int(current_user.id):
        flash('不能删除自己', 'warning')
        return redirect(url_for('admin.users'))
    if user['is_admin'] and _admin_count() <= 1:
        flash('不能删除最后一位管理员', 'warning')
        return redirect(url_for('admin.users'))
    try:
        db.delete_user(uid)
    except db.UserHasContractsError as e:
        flash(f'不能删除用户「{user["username"]}」：其名下还有 {e.n} 个合同，'
              f'请先在合同中转移经办人', 'warning')
        return redirect(url_for('admin.users'))
    except sqlite3.IntegrityError:
        flash(f'删除失败：用户「{user["username"]}」存在关联数据', 'danger')
        return redirect(url_for('admin.users'))
    flash(f'已删除用户：{user["username"]}', 'success')
    log_action('删除用户', target=user['username'])
    return redirect(url_for('admin.users'))


@bp.route('/admin/backup')
@admin_required
def backup():
    """一键备份数据库并下载（bill_backup_YYYYMMDD.db）。"""
    try:
        path = db.backup_db()
    except Exception as e:
        flash(f'备份失败：{e}', 'danger')
        return redirect(url_for('admin.users'))
    log_action('备份数据库')
    return send_file(path, as_attachment=True,
                     download_name=os.path.basename(path), mimetype='application/x-sqlite3')


def _admin_count():
    return sum(1 for u in db.list_users() if u['is_admin'])


# ============ 用户批量导入（Excel）============
_TRUE_WORDS = {'是', 'y', 'Y', '1', 'true', 'TRUE', 'True', '管理员', 'admin'}


def _import_col(header, keywords):
    """按表头关键字定位列号（含「用户名」「账号」等模糊匹配）。"""
    for i, h in enumerate(header):
        if h and any(k in str(h) for k in keywords):
            return i
    return None


@bp.route('/admin/users/import', methods=['POST'])
@admin_required
def users_import():
    """批量导入用户：Excel 列「用户名 / 显示名 / 是否管理员」。
    初始密码统一为 IMPORT_DEFAULT_PASSWORD（可用 BILLHUB_IMPORT_PASSWORD 覆盖），
    导入用户首次登录强制改密；重复用户名跳过。"""
    f = request.files.get('file')
    if not f or not f.filename:
        flash('请选择要导入的 Excel 文件（.xlsx）', 'danger')
        return redirect(url_for('admin.users'))
    if not f.filename.lower().endswith('.xlsx'):
        flash('仅支持 .xlsx 格式的 Excel 文件', 'danger')
        return redirect(url_for('admin.users'))
    default_pwd = current_app.config['IMPORT_DEFAULT_PASSWORD']
    err = password_error(default_pwd)
    if err:
        flash(f'导入初始密码不符合强度要求（{err}），请检查 BILLHUB_IMPORT_PASSWORD 配置', 'danger')
        return redirect(url_for('admin.users'))
    try:
        wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
    except Exception as e:
        flash(f'Excel 解析失败：{e}', 'danger')
        return redirect(url_for('admin.users'))
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    header = next(rows, None)
    if not header:
        wb.close()
        flash('Excel 内容为空，请按模板填写后再导入', 'danger')
        return redirect(url_for('admin.users'))
    ci_user = _import_col(header, ('用户名', '账号'))
    ci_disp = _import_col(header, ('显示名', '姓名'))
    ci_admin = _import_col(header, ('管理员', '角色'))
    if ci_user is None:
        wb.close()
        flash('未找到「用户名」列（表头需含：用户名 / 显示名 / 是否管理员，可先下载导入模板）',
              'danger')
        return redirect(url_for('admin.users'))
    pwd_hash = hash_password(default_pwd)
    created, skipped = [], []
    seen = set()
    for row in rows:
        if row is None:
            continue
        username = str(row[ci_user] or '').strip() if ci_user < len(row) else ''
        if not username:
            continue
        disp = str(row[ci_disp] or '').strip() if (ci_disp is not None and ci_disp < len(row)) else ''
        admin_raw = str(row[ci_admin] or '').strip() if (ci_admin is not None and ci_admin < len(row)) else ''
        is_admin = admin_raw in _TRUE_WORDS
        if username in seen or db.get_user_by_username(username):
            skipped.append(username)
            continue
        seen.add(username)
        if db.create_user(username, pwd_hash, display_name=disp,
                          is_admin=is_admin, must_change_password=1) is None:
            skipped.append(username)
        else:
            created.append(username)
    wb.close()
    msg = f'导入完成：新增 {len(created)} 人'
    if skipped:
        msg += f'，跳过 {len(skipped)} 人（重复用户名：{"、".join(skipped)}）'
    flash(msg, 'success' if created else 'warning')
    if created:
        flash(f'导入用户初始密码为默认密码（首登强制修改），请妥善分发给 {len(created)} 位新用户', 'info')
    log_action('批量导入用户', detail=msg)
    return redirect(url_for('admin.users'))


@bp.route('/admin/users/import/template')
@admin_required
def users_import_template():
    """下载用户导入模板（用户名 / 显示名 / 是否管理员）。"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '用户'
    ws.append(['用户名', '显示名', '是否管理员'])
    ws.append(['zhangsan', '张三', '否'])
    ws.append(['lisi', '李四', '是'])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name='用户导入模板.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ============ OCR 规则配置（可配置化，保存后立即生效）============
def _lines(value):
    return [ln.strip() for ln in (value or '').splitlines() if ln.strip()]


_ALT_KEYS = ('payee_pattern', 'bank_name_pattern', 'bank_account_pattern')
_ALT_FIELD_NAMES = {
    'payee_pattern': '收款单位', 'bank_name_pattern': '开户银行',
    'bank_account_pattern': '银行账号',
}
_RESETTABLE_KEYS = ('amount_keywords', 'start_date_keywords', 'end_date_keywords',
                    'payee_pattern', 'bank_name_pattern', 'bank_account_pattern',
                    'plan_kw_map')
# 期数下拉语义（与识别排序逻辑一致：9=末期、10=质保期）
PLAN_SEQ_OPTIONS = [('1', '第1期'), ('2', '第2期'), ('3', '第3期'), ('4', '第4期'),
                    ('9', '末期（尾款 / 余款）'), ('10', '质保期（质保金）')]


def _ocr_view(saved):
    """构建规则页视图模型：每字段的生效值 / 默认值 / 是否自定义 / 编辑模式。"""
    rules = ocr.merge_contract_rules(saved)
    view = {'fields': {}, 'plan': {'rows': [], 'custom': False}}
    for key in ('amount_keywords', 'start_date_keywords', 'end_date_keywords'):
        view['fields'][key] = {
            'current': rules[key],
            'default': ocr.DEFAULT_CONTRACT_RULES[key],
            'custom': key in saved and bool(saved.get(key)),
        }
    for key in _ALT_KEYS:
        cur = rules[key]
        kws = ocr.split_alt_prefix(cur, ocr.ALT_TAILS[key])
        view['fields'][key] = {
            'current': cur,
            'keywords': kws or [],
            'mode': 'simple' if kws is not None else 'advanced',
            'custom': key in saved and bool(saved.get(key)),
            'default': ocr.DEFAULT_CONTRACT_RULES[key],
        }
    plan = rules['plan_kw_map']
    view['plan'] = {
        'rows': [{'kw': k, 'seq': str(v)} for k, v in plan.items()],
        'custom': 'plan_kw_map' in saved and bool(saved.get('plan_kw_map')),
    }
    return view


def _overrides_from_form(form):
    """从表单构建 overrides（保存与测试器共用）。返回 (overrides, error)。"""
    overrides = {
        'amount_keywords': _lines(form.get('amount_keywords')),
        'start_date_keywords': _lines(form.get('start_date_keywords')),
        'end_date_keywords': _lines(form.get('end_date_keywords')),
        'plan_kw_map': {},
    }
    for key in _ALT_KEYS:
        kw_name = f'{key.split("_pattern")[0]}_keywords'
        if form.get(f'{key}_mode', 'simple') == 'advanced':
            overrides[key] = form.get(key, '').strip()
        else:
            kws = _lines(form.get(kw_name))
            overrides[key] = ocr.pattern_from_keywords(kws, ocr.ALT_TAILS[key]) if kws else ''
    for key in _ALT_KEYS:
        p = overrides[key]
        if p:
            try:
                re.compile(p)
            except re.error as e:
                return None, f'正则不合法（{_ALT_FIELD_NAMES[key]}）：{e}'
    plan_kws = form.getlist('plan_kw')
    plan_seqs = form.getlist('plan_seq')
    for kw, seq in zip(plan_kws, plan_seqs):
        kw = (kw or '').strip()
        if not kw:
            continue
        try:
            overrides['plan_kw_map'][kw] = int(seq)
        except (TypeError, ValueError):
            return None, f'「{kw}」的期数选择无效，请重新选择'
    overrides = {k: v for k, v in overrides.items() if v}
    return overrides, None


@bp.route('/admin/ocr-rules')
@admin_required
def ocr_rules():
    saved = {}
    raw = db.get_setting('ocr_rules')
    if raw:
        try:
            saved = json.loads(raw) or {}
        except ValueError:
            saved = {}
    return render_template('admin/ocr_rules.html', view=_ocr_view(saved),
                           seq_options=PLAN_SEQ_OPTIONS)


@bp.route('/admin/ocr-rules', methods=['POST'])
@admin_required
def ocr_rules_save():
    overrides, err = _overrides_from_form(request.form)
    if err:
        flash(err, 'danger')
        return redirect(url_for('admin.ocr_rules'))
    db.set_settings({'ocr_rules': json.dumps(overrides, ensure_ascii=False)})
    flash('OCR 规则已保存，重新识别立即生效', 'success')
    log_action('保存 OCR 规则')
    return redirect(url_for('admin.ocr_rules'))


@bp.route('/admin/ocr-rules/reset', methods=['POST'])
@admin_required
def ocr_rules_reset():
    db.delete_settings(['ocr_rules'])
    flash('已恢复出厂默认 OCR 规则', 'success')
    log_action('恢复默认 OCR 规则')
    return redirect(url_for('admin.ocr_rules'))


@bp.route('/admin/ocr-rules/reset-field', methods=['POST'])
@admin_required
def ocr_rules_reset_field():
    """单字段恢复出厂默认（其余字段保留）。"""
    key = request.form.get('field', '').strip()
    if key not in _RESETTABLE_KEYS:
        flash('未知的规则字段', 'danger')
        return redirect(url_for('admin.ocr_rules'))
    saved = {}
    raw = db.get_setting('ocr_rules')
    if raw:
        try:
            saved = json.loads(raw) or {}
        except ValueError:
            saved = {}
    saved.pop(key, None)
    if saved:
        db.set_settings({'ocr_rules': json.dumps(saved, ensure_ascii=False)})
    else:
        db.delete_settings(['ocr_rules'])
    flash('该字段已恢复出厂默认', 'success')
    log_action('恢复默认 OCR 规则', detail=f'单字段：{key}')
    return redirect(url_for('admin.ocr_rules'))


@bp.route('/admin/ocr-rules/test', methods=['POST'])
@admin_required
def ocr_rules_test():
    """规则测试器：用表单草稿规则（未保存）解析文本，返回各字段结果 JSON。"""
    text = request.form.get('text', '').strip()
    if not text:
        return jsonify({'ok': False, 'error': '请输入要测试的合同文本片段'})
    overrides, err = _overrides_from_form(request.form)
    if err:
        return jsonify({'ok': False, 'error': err})
    try:
        data = ocr.extract_from_text(text, overrides)
    except Exception as e:
        return jsonify({'ok': False, 'error': f'解析失败：{e}'})
    if not data:
        return jsonify({'ok': True, 'fields': {}, 'message': '未识别到有效信息'})
    return jsonify({'ok': True, 'fields': data})


# ============ LDAP/AD 配置（可视化，保存后立即生效）============
@bp.route('/admin/ldap')
@admin_required
def ldap_config():
    saved = db.get_settings()
    return render_template('admin/ldap.html', cfg=saved)


@bp.route('/admin/ldap', methods=['POST'])
@admin_required
def ldap_config_save():
    items = {
        'ldap_enabled': '1' if request.form.get('ldap_enabled') == 'on' else '0',
        'ldap_uri': request.form.get('ldap_uri', '').strip(),
        'ldap_base_dn': request.form.get('ldap_base_dn', '').strip(),
        'ldap_bind_dn': request.form.get('ldap_bind_dn', '').strip(),
        'ldap_user_dn_template': request.form.get('ldap_user_dn_template', '').strip(),
        'ldap_search_filter': request.form.get('ldap_search_filter', '').strip(),
    }
    bind_pwd = request.form.get('ldap_bind_password', '')
    if bind_pwd:  # 留空保持原值
        items['ldap_bind_password'] = bind_pwd
    db.set_settings(items)
    flash('LDAP 配置已保存，立即生效', 'success')
    log_action('保存 LDAP 配置')
    return redirect(url_for('admin.ldap_config'))


@bp.route('/admin/ldap/test', methods=['POST'])
@admin_required
def ldap_config_test():
    """测试连接 / 测试认证：使用表单中当前填写的参数（未保存也能测）。"""
    from web.config import effective_ldap_config
    from web.ldap_auth import test_connection
    cfg = effective_ldap_config(current_app)
    form_map = {'ldap_uri': 'LDAP_URI', 'ldap_base_dn': 'LDAP_BASE_DN',
                'ldap_bind_dn': 'LDAP_BIND_DN',
                'ldap_user_dn_template': 'LDAP_USER_DN_TEMPLATE',
                'ldap_search_filter': 'LDAP_SEARCH_FILTER'}
    for fk, ck in form_map.items():
        v = request.form.get(fk, '').strip()
        if v:
            cfg[ck] = v
    cfg['LDAP_ENABLED'] = request.form.get('ldap_enabled') == 'on'
    pwd = request.form.get('ldap_bind_password', '')
    if pwd:
        cfg['LDAP_BIND_PASSWORD'] = pwd
    ok, msg = test_connection(cfg,
                              request.form.get('test_username', '').strip(),
                              request.form.get('test_password', ''))
    return jsonify({'ok': ok, 'message': msg})


# ============ 审计日志（仅管理员，只读）============
AUDIT_PER_PAGE = 20


@bp.route('/admin/audit-logs')
@admin_required
def audit_logs():
    username = request.args.get('username', '').strip()
    action = request.args.get('action', '').strip()
    start_date = request.args.get('start_date', '').strip()
    end_date = request.args.get('end_date', '').strip()
    page = max(1, request.args.get('page', 1, type=int))
    rows, total = db.list_audit_logs(page=page, per_page=AUDIT_PER_PAGE,
                                     username=username or None,
                                     action=action or None,
                                     start_date=start_date or None,
                                     end_date=end_date or None)
    total_pages = max(1, (total + AUDIT_PER_PAGE - 1) // AUDIT_PER_PAGE)
    page = min(page, total_pages)
    rows, _ = db.list_audit_logs(page=page, per_page=AUDIT_PER_PAGE,
                                 username=username or None,
                                 action=action or None,
                                 start_date=start_date or None,
                                 end_date=end_date or None)
    return render_template('admin/audit_logs.html', rows=rows, total=total,
                           page=page, per_page=AUDIT_PER_PAGE, total_pages=total_pages,
                           username=username, action=action,
                           start_date=start_date, end_date=end_date,
                           actions=db.list_audit_actions())
